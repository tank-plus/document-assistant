# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# # 创建工作簿
# wb = Workbook()

# # 选择活动工作表
# ws = wb.active

# # 设置标题和公司信息
# ws['A1'] = 'NINGBO HARBOR IMPORT & EXPORT CO.,LTD'
# ws['A3'] = 'Bizhu Chenjia Village, Xikou Town, Fenghua City, Zhejiang Province, China'
# ws['A4'] = 'Tel: +86-574-88807585'
# ws['A5'] = 'Fax: +86-574-88806858'

# # 合并单元格
# ws.merge_cells('A1:D1')
# ws.merge_cells('A3:D3')

# # 格式化标题和公司信息
# title_font = Font(bold=True, size=14)
# title_alignment = Alignment(horizontal='center', vertical='center')
# title_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                       top=Side(style='thin'), bottom=Side(style='thin'))
# title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# for row in ws['1:5']:
#     for cell in row:
#         cell.font = title_font
#         cell.alignment = title_alignment
#         cell.border = title_border
#         cell.fill = title_fill

# # 客户信息
# ws['A7'] = 'To:'
# ws['B7'] = 'Real Flame Company Inc.'
# ws['B8'] = '7800 Northwestern Avenue, Racine, WI 53406'
# ws['B9'] = 'Tel/Fax: 800 6541704'
# ws['A10'] = 'P/I NO.: NBHB/RFC2425'
# ws['B10'] = 'PO NO.: 607132'
# ws['A11'] = 'Date: 2024.5.3'

# # 表格头部
# headers = ["Description", "Model No.", "Quantity", "Unit Price", "Total Amount"]
# ws.append(headers)

# # 设置头部样式
# for cell in ws[ws.max_row-1]:
#     cell.font = Font(bold=True)
#     cell.border = Border(left=Side(style='medium'), right=Side(style='medium'),
#                          top=Side(style='medium'), bottom=Side(style='medium'))
#     cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

# # 表格数据
# items = [
#     ["FIREPLACE MANTEL", "1290-X-AGR", "84PCS/168CTNS", 155.8, 13087.2],
#     ["FIREPLACE MANTEL", "1290-X-W", "84PCS/168CTNS", 150.1, 12608.4],
# ]

# # 填充数据
# for item in items:
#     ws.append(item)

# # 总计
# ws['A14'] = 'TOTAL:'
# ws['C14'] = '168PCS/336CTNS'
# ws['D14'] = 150.1
# ws['E14'] = 25695.6

# # 总计说明
# total_text = "SAY US DOLLARS TWENTY FIVE THOUSAND SIX HUNDRED AND NINETY FIVE CENTS SIXTY ONLY."
# ws['A15'] = total_text
# ws['A16'] = total_text.upper()

# # 条款和条件
# ws['A18'] = "TERMS OF PAYMENT:"
# ws['A19'] = "NO DEPOSIT, T/T WITHIN 21DAYS AFTER SHIPMENT"
# ws['A20'] = "TIME OF SHIPMENT:     JUN 11TH, 2024"
# ws['A21'] = "PORT OF LOADING:      NINGBO, CHINA"
# ws['A22'] = "PORT OF DESTINATION:  SACRAMENTO, CA, USA"

# # 银行信息
# bank_info = [
#     "BENEFICIARY’S BANK:",
#     "Account with institution: ZHEJIANG RURAL CREDIT COOPERATIVE UNION, HANGZHOU, CHINA",
#     "SWIFT BIC:ZJRCCN2N",
#     "NO.660 QIUTAO ROAD, HANGZHOU, ZHEJIAHG PROVINCE, CHINA, 310016",
#     "Beneficiary: NINGBO HARBOR IMPORT & EXPORT CO.,LTD",
#     "Account Number: 201000134920721",
# ]

# ws.append(bank_info)

# # 保存工作簿
# wb.save('ProformaInvoice_complete_example.xlsx')

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# 创建一个新的工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active

ws.title = "PI"

# 定义字体和对齐方式
font18_bold = Font(size=18, bold=True)
font16_black = Font(size=16, name="黑体")
font12 = Font(size=12)
font12_bold = Font(size=12, bold=True)
font11 = Font(size=11)

align_center = Alignment(horizontal='center', vertical='center')

# 定义边框
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# 第一行
ws.merge_cells('A1:E1')
ws['A1'] = '第一行内容'
ws['A1'].font = font18_bold
ws['A1'].alignment = align_center

# 第二行
ws.merge_cells('A2:E2')
ws['A2'] = '第二行内容'
ws['A2'].font = font16_black
ws['A2'].alignment = align_center

# 第三行
ws.merge_cells('A3:E3')
ws['A3'] = '第三行内容'
ws['A3'].font = font12
ws['A3'].alignment = align_center

# 第四行
ws.merge_cells('A4:E4')
ws['A4'] = '第四行内容'
ws['A4'].font = font12
ws['A4'].alignment = align_center
ws['A4'].border = Border(bottom=Side(style='thin'))

# 空一行

# 第六行
ws.merge_cells('A6:E6')
ws['A6'] = '第六行内容'
ws['A6'].font = font18_bold
ws['A6'].alignment = align_center

# 空一行

# 第八行
ws['A8'] = 'A8内容'
ws['A8'].font = font11

ws['D8'] = 'D8内容'
ws['D8'].font = font11

ws['E8'] = 'E8内容'
ws['E8'].font = font11

# 第九行
ws['A9'] = 'A9内容'
ws['A9'].font = font11

ws['D9'] = 'D9内容'
ws['D9'].font = font11

ws['E9'] = 'E9内容'
ws['E9'].font = font11

# 第十行
ws['A10'] = 'A10内容'
ws['A10'].font = font11

ws['D10'] = 'D10内容'
ws['D10'].font = font11

ws['E10'] = 'E10内容'
ws['E10'].font = font11

# 三行空行

# 第十四行和第十五行
ws.merge_cells('A14:A15')
ws['A14'] = 'A14:A15内容'
ws['A14'].font = font12_bold
ws['A14'].alignment = align_center
ws['A14'].border = thin_border

ws.merge_cells('B14:B15')
ws['B14'] = 'B14:B15内容'
ws['B14'].font = font12_bold
ws['B14'].alignment = align_center
ws['B14'].border = thin_border

ws.merge_cells('C14:C15')
ws['C14'] = 'C14:C15内容'
ws['C14'].font = font12_bold
ws['C14'].alignment = align_center
ws['C14'].border = thin_border

ws['D14'] = 'D14内容'
ws['D14'].border = thin_border

ws['D15'] = 'D15内容'
ws['D15'].border = thin_border

ws.merge_cells('E14:E15')
ws['E14'] = 'E14:E15内容'
ws['E14'].font = font12_bold
ws['E14'].alignment = align_center
ws['E14'].border = thin_border

# 遍历数据部分
start_row = 16
data = [
    {'category': '类别1', 'sum': 100},
    {'category': '类别2', 'sum': 200},
]

for item in data:
    ws.merge_cells(f'A{start_row}:B{start_row}')
    ws[f'A{start_row}'] = item['category']
    ws[f'A{start_row}'].font = font12_bold
    ws[f'A{start_row}'].alignment = align_center
    ws[f'A{start_row}'].border = thin_border

    ws[f'E{start_row}'] = item['sum']
    ws[f'E{start_row}'].number_format = '"$"#,##0.00'
    ws[f'E{start_row}'].font = font12_bold
    ws[f'E{start_row}'].alignment = align_center
    ws[f'E{start_row}'].border = thin_border

    ws.merge_cells(f'B{start_row + 1}:E{start_row + 1}')
    ws[f'B{start_row + 1}'] = '内容'
    ws[f'B{start_row + 1}'].font = font12
    ws[f'B{start_row + 1}'].alignment = align_center
    ws[f'B{start_row + 1}'].border = thin_border

    start_row += 2

    for i in range(5):
        ws.merge_cells(f'B{start_row}:E{start_row}')
        ws[f'B{start_row}'] = '内容'
        ws[f'B{start_row}'].font = font12
        ws[f'B{start_row}'].alignment = align_center
        ws[f'B{start_row}'].border = thin_border
        start_row += 1

# "The Seller" 放图片
img = Image("logo.png")
ws.add_image(img, 'B48')

# 保存文件
wb.save('formatted_excel.xlsx')
