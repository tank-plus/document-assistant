from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side
from core.util import *
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from model.order import Order
import os

# 定义边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 定义字体
font24_bold = Font(size=24, bold=True, name="Times New Roman")
font12_bold = Font(size=12, bold=True, name="Times New Roman")
font10_common = Font(size=10, name="Times New Roman")
font11_common = Font(size=11, name="Times New Roman")
font12_underline = Font(size=12, underline='single', name="Times New Roman")
font10_bold_underline = Font(size=10, bold=True, underline='single', name="Times New Roman")


class OrderPIProcessor:
    def __init__(self, order: Order, template="./template/order_pi_template.xlsx"):
        self.order = order
        if not os.path.exists("./static/pi_confirm"):
            os.makedirs("./static/pi_confirm")
        self.file_name = f"./static/pi_confirm/{self.order.po_num}.xlsx"
        self.template = template
        
    @property
    def target_file_name(self):
        return self.file_name.replace("./","/")
    
    def transfer_order_data(self):
        order_real_data = {}
        total_amount = 0.0
        total_pcs = 0
        total_ctns = 0
        for order_detail in self.order.order_details:
            total_amount += order_detail.total_price
            total_pcs += order_detail.qty
            total_ctns += order_detail.ctns
            pi_category = order_detail.pi_category
            if pi_category not in order_real_data:
                order_real_data[pi_category] = []
            
            order_real_data[pi_category].append([pi_category,
                                                 order_detail.product_id, 
                                                 f"{order_detail.qty}PCS/{order_detail.ctns}CTNS", 
                                                 order_detail.unit_price, 
                                                 order_detail.total_price])
        for part_detail in self.order.part_details:
            total_amount += part_detail.total_price
            total_pcs += part_detail.qty
            total_ctns += part_detail.ctns
            pi_category = part_detail.pi_category
            if pi_category not in order_real_data:
                order_real_data[pi_category] = []
            
            order_real_data[pi_category].append([pi_category,
                                                 part_detail.product_id, 
                                                 f"{part_detail.qty}PCS/{part_detail.ctns}CTNS", 
                                                 part_detail.unit_price, 
                                                 part_detail.total_price])
        return order_real_data, total_amount, total_pcs, total_ctns

    def process(self):
        # 加载Excel文件
        workbook = load_workbook(filename=self.template, data_only=False)  # 替换为你的Excel文件名
        # 遍历所有sheet
        for sheet_name in workbook.sheetnames:
            description_row = None
            print(f"正在处理Sheet: {sheet_name}")
            if sheet_name == 'PI':
                data_rows = 0
                order_real_data, total_amount, total_pcs, total_ctns = self.transfer_order_data()

                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        # 查找Description所在的行
                        if cell.value == "Description":
                            description_row = cell.row

                        if cell.value and "The Seller:" in str(cell.value):
                            picture_row = cell.row
                            print(f"picture_row:{picture_row}")

                if description_row:
                    for key, value in order_real_data.items():
                        data_rows += len(value)

                    print(f"data_rows:{data_rows}")

                    sheet.insert_rows(description_row + 2, data_rows)  # 插入行

                # 拆分数据区域内的合并单元格
                list = []
                for cell_range in sheet.merged_cells:
                    list.append(cell_range)

                for cell_range in list:
                    try:
                        start_row = cell_range.min_row
                        end_row = cell_range.max_row
                        if start_row >= description_row + 2 and end_row <= description_row + 2 + data_rows:
                            print(f"{cell_range}")
                            sheet.unmerge_cells(str(cell_range))
                    except Exception as e:
                        print(e)

                start_row = description_row + 2
                catogery_row_pair = []
                for key, records in order_real_data.items():
                    for i, data_record in enumerate(records, start=start_row):
                        for j, value in enumerate(data_record, start=1):
                            cell = sheet.cell(row=i, column=j)
                            cell.value = value
                            cell.border = thin_border
                            cell.number_format = '$#,##0.00'
                            cell.font = font11_common
                            cell.alignment = Alignment(vertical='center')
                            
                    catogery_row_pair.append((start_row, start_row + len(records) - 1))
                    start_row += len(records)
                    
                

                total_row = description_row + 2 + data_rows
            

                print(f"total_row:{total_row}")

                total_cell1 = sheet.cell(row=total_row, column=1)
                total_cell1.value = 'TOTAL:'
                total_cell1.border = thin_border
                total_cell1.font = font12_bold
                total_cell1.alignment = Alignment(horizontal='right', vertical='center')
                
                
                total_cell3 = sheet.cell(row=total_row, column=3)
                # TODO 汇总PCS/CTNS
                total_cell3.value = f'{total_pcs}PCS/{total_ctns}CTNS'
                total_cell3.border = thin_border
                total_cell3.font = font10_bold_underline
                total_cell3.alignment = Alignment(horizontal='left', vertical='center')

                total_cell5 = sheet.cell(row=total_row, column=5)
                total_cell5.value = total_amount
                total_cell5.border = thin_border
                total_cell5.number_format = '$#,##0.00'
                total_cell5.font = font10_bold_underline

                sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)

                sheet.cell(row=total_row + 1, column=2, value=number_to_words(total_amount))
                sheet.merge_cells(f"B{total_row + 1}:E{total_row + 1}")
                cell = sheet[f'B{total_row + 2}']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                # 类别相关的列合并
                for (start, end) in catogery_row_pair:
                    sheet.merge_cells(f"A{start}:A{end}")
                    cell_new = sheet[f"A{start}"]
                    cell_new.alignment = Alignment(horizontal='center', vertical='center')

                rich_text_a8 = CellRichText('To:', TextBlock(InlineFont(rFont='Times New Roman', u='single', sz=12), self.order.customer.customer_id))
                rich_text_a8.font = Font(size=14, bold=True, name="Times New Roman")
                sheet["A8"] = rich_text_a8

                rich_text_a9 = CellRichText('Add:', TextBlock(InlineFont(rFont='Times New Roman', u='single', sz=12),self.order.customer.address))
                rich_text_a9.font = Font(size=14, bold=True, name="Times New Roman")
                sheet["A9"] = rich_text_a9

                rich_text_a10 = CellRichText('Tel/Fax:',TextBlock(InlineFont(rFont='Times New Roman', u='single', sz=12),self.order.customer.tel_fax))
                rich_text_a10.font = Font(size=14, bold=True, name="Times New Roman")
                sheet["A10"] = rich_text_a10

                img = Image('logo.png')  # 替换为图片文件的路径

                # 设置图片的宽度和高度
                img.width = 250  # 设置宽度为100像素
                img.height = 125  # 设置高度为100像素

                sheet.add_image(img, f'B{picture_row + len(order_real_data) + 1}')
                
                
                # replace_data
                data = {}
                data["{{pi_num}}"] = self.order.pi_num
                data["{{po_num}}"] = self.order.po_num
                data["{{po_date}}"] = self.order.po_date.strftime("%Y-%m-%d")
                data["{{customer_name}}"] = "To: " + self.order.customer.customer_id
                data["{{customer_address}}"] = "Add: " + self.order.customer.address
                data["{{customer_tel_fax}}"] = "Tel/Fax: " + self.order.customer.tel_fax
                data["{{payment_method}}"] = self.order.payment_method
                data["{{delivery_date}}"] = self.order.delivery_date
                data["{{port_of_loading}}"] = self.order.port_of_loading
                data["{{port_of_destination}}"] = self.order.port_of_destination
                # 遍历sheet中的所有行和单元格
                for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
                    for cell in row:
                        if cell.value:
                            for placeholder, actual_value in data.items():
                                if placeholder in str(cell.value):
                                    cell.value = str(cell.value).replace(placeholder, actual_value)
                                    
                                    
                                    
                # 最后一个单元格的金额样式调整
                sheet[f"E{total_row-1}"].alignment = Alignment(horizontal='right', vertical='center')
                sheet["A1"].font = font24_bold
                sheet["E8"].font = font12_underline
                sheet["E9"].font = font12_underline
                sheet["E10"].font = font12_underline
                sheet.row_dimensions[total_row].height =sheet.row_dimensions[total_row-1].height
                
                        

        # 保存Excel文件
        workbook.save(self.file_name)


if __name__ == "__main__":
    data = {
        "{{pi_num}}": "NBHB/RFC2331",
        "{{po_num}}": "606234",
        "{{po_date}}": "2021-01-01",
        "{{customer_name}}": "Real Flame Company Inc.  ",
        "{{customer_address}}": "7800 Northwestern Avenue,Racine, WI 53406",
        "{{customer_tel_fax}}": "800 6541704",
        "{{payment_method}}": "NO DEPOSIT. T/T WITHIN 21DAYS AFTER SHIPMENT",
        "{{delivery_date}}": "AUG 18TH, 2023",
        "{{port_of_loading}}": "NINGBO, CHINA",
        "{{port_of_destination}}": "MONTREAL, CANADA",
        "mantel": [
            ["FIREPLACE MANTEL", "13058-X-VBM", "120PCS/120CTNS", "$111.2", "$13,344.00"],
            ["FIREPLACE MANTEL", "8700-X-CHBW", "109PCS/218CTNS", "$154.90", "$16,884.10 "],
            ["FIREPLACE MANTEL", "8701-X-CHBW", "110PCS/110CTNS", "$154.90", "$16,884.10 "],
            ["FIREPLACE MANTEL", "13058-X-VBM", "120PCS/120CTNS", "$111.2", "$13,344.00"],
            ["FIREPLACE MANTEL", "8700-X-CHBW", "109PCS/218CTNS", "$154.90", "$16,884.10 "],
            ["FIREPLACE MANTEL", "8701-X-CHBW", "110PCS/110CTNS", "$154.90", "$16,884.10 "],
            ["FIREPLACE MANTEL", "13058-X-VBM", "120PCS/120CTNS", "$111.2", "$13,344.00"],
            ["FIREPLACE MANTEL", "8700-X-CHBW", "109PCS/218CTNS", "$154.90", "$16,884.10 "],
            ["FIREPLACE MANTEL", "8701-X-CHBW", "110PCS/110CTNS", "$154.90", "$16,884.10 "],
            ["FIREPLACE MANTEL", "13058-X-VBM", "120PCS/120CTNS", "$111.2", "$13,344.00"],
            ["FIREPLACE MANTEL", "8700-X-CHBW", "109PCS/218CTNS", "$154.90", "$16,884.10 "],
            ["FIREPLACE MANTEL", "8701-X-CHBW", "110PCS/110CTNS", "$154.90", "$16,884.10 "]
        ]
    }

    processor = OrderPIProcessor(data, "output.xlsx")
    processor.process()
