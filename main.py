from openpyxl import load_workbook
import os


def list_all_files(directory, suffix=".xlsx", with_order=True):
    list = []
    csv_files = [file for file in os.listdir(directory) if file.endswith(suffix)]
    
    if with_order:
        csv_files = sorted(csv_files)
    
    for file in csv_files:
        list.append(os.path.join(directory, file))
    return list


# 加载基础数据
BASIC = {}
FIREPLACE_MANTEL = []
FIREPLACE_MANTEL_TOP = []
CUSTOMER = []

workbook = load_workbook(filename='数据.xlsx', data_only=True)  # 


# TODO 将汇总数据塞进BASIC，用于替换
for sheet_name in workbook.sheetnames:
    # print(f"正在处理Sheet: {sheet_name}")
    
    
    sheet = workbook[sheet_name]

    if "基本信息" == sheet_name:
        # 遍历sheet中的所有行和单元格
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
            BASIC["{{" + row[0].value + "}}"] = row[1].value
        # print(BASIC)
        
    if "FIREPLACE_MANTEL" == sheet_name:
        index = 0
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
            record = []
            for cell in row:
                record.append(cell.value)
            FIREPLACE_MANTEL.append(record)
            index = index + 1
        # print(FIREPLACE_MANTEL)
        
        
    if "FIREPLACE_MANTEL_TOP" == sheet_name:
        index = 0
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
            record = []
            for cell in row:
                record.append(cell.value)
            FIREPLACE_MANTEL_TOP.append(record)
            index = index + 1
        # print(FIREPLACE_MANTEL_TOP)
        
    if "客户信息" == sheet_name:
        index = 0
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
            record = []
            for cell in row:
                record.append(cell.value)
            CUSTOMER.append(record)
            index = index + 1
        # print(CUSTOMER)




# 替换模板内的数据并生成新文件
# 加载Excel文件
templates = list_all_files('./template', with_order=False)

for template in templates:
    workbook = load_workbook(filename=template)  # 替换为你的Excel文件名

    # 遍历所有sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 遍历sheet中的所有行和单元格
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if cell.value:
                    for placeholder, actual_value in BASIC.items():
                        if placeholder in str(cell.value):
                            cell.value = str(cell.value).replace(placeholder, actual_value)

    # 保存替换后的文件
    workbook.save(filename=f'output/{os.path.basename(template)}')  # 输出文件名可以根据需要进行修改

print("数据处理完毕，数据已替换。")
