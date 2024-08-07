from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side
from core.util import *
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

BASIC = {}
FIREPLACE_MANTEL = []
FIREPLACE_MANTEL_TOP = []
CUSTOMER = []

# 定义边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 定义字体
font12_bold = Font(size=12, bold=True, name="Times New Roman")
font10_common = Font(size=10, name="Times New Roman")
font10_bold_underline = Font(size=10, bold=True, underline='single', name="Times New Roman")



# 加载Excel文件
workbook = load_workbook(filename='template/order_pi_template.xlsx', data_only=False)  # 替换为你的Excel文件名

# 遍历所有sheet
for sheet_name in workbook.sheetnames:

    description_row = None
    print(f"正在处理Sheet: {sheet_name}")
    

    if sheet_name == 'PI':
    
        sheet = workbook[sheet_name]

        
        for row in sheet.iter_rows():
            for cell in row:
                # 查找Description所在的行
                if cell.value == "Description":
                    description_row = cell.row
                
                if cell.value and "The Seller:" in str(cell.value):
                    picture_row = cell.row
                    print(f"picture_row:{picture_row}")
            
        
        if description_row:
            new_data =[
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00],
                ['FIREPLACE MANTEL', '13058-X-VBM', '120PCS/120CTNS',111.20,13344.00]
            ]
            sheet.insert_rows(description_row + 2, len(new_data))  # 插入行

        
        # 拆分数据区域内的合并单元格
        list=[]
        for cell_range in sheet.merged_cells:
            list.append(cell_range)

        for cell_range in list:
            try:
                start_row = cell_range.min_row
                end_row = cell_range.max_row
                if start_row >= description_row + 2 and end_row <= description_row + 2 + len(new_data):
                    print(f"{cell_range}")
                    sheet.unmerge_cells(str(cell_range))
            except Exception as e:
                print(e)
                
        for i, data_row in enumerate(new_data, start=description_row + 2):
            for j, value in enumerate(data_row, start=1):
                cell = sheet.cell(row=i, column=j)
                cell.value = value
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                cell.font = font10_common
            
        total_row = description_row + 2 + len(new_data)
        
        
        print(f"total_row:{total_row}")
        
        
        total_cell1 = sheet.cell(row=total_row, column=1)
        total_cell1.value = 'TOTAL:'
        total_cell1.border = thin_border
        total_cell1.font = font12_bold
        total_cell1.alignment = Alignment(horizontal='right', vertical='center')
        
        # total_cell3 = sheet.cell(row=total_row, column=3)
        # total_cell3.value = 120
        # total_cell3.number_format = '$#,##0.00'
        # total_cell3.font = font10_bold_underline
        # total_cell3.border = thin_border
        
        total_cell5 = sheet.cell(row=total_row, column=5)
        total_cell5.value = 13344.00
        total_cell5.border = thin_border
        total_cell5.number_format = '$#,##0.00'
        total_cell5.font = font10_bold_underline
       
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)

        # TODO 金额相关的列合并
        amount = 13344.22
        sheet.cell(row=total_row + 1, column=2, value=number_to_words(amount))
        sheet.merge_cells(f"B{total_row + 1}:E{total_row + 1}")
        cell = sheet[f'B{total_row + 2}']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        
        
        # TODO 类别相关的列合并
        sheet.merge_cells(f"A{description_row + 2}:A{description_row + 1+ len(new_data)}")
        cell_new = sheet[f"A{description_row + 2}"]
        cell_new.alignment = Alignment(horizontal='center', vertical='center')


        rich_text_a8 = CellRichText('To:', TextBlock(InlineFont(rFont='Times New Roman',u='single'), 'Real Flame Company Inc.'))
        rich_text_a8.font = Font(size=12, bold=True, name="Times New Roman")
        sheet["A8"] = rich_text_a8      

        rich_text_a9 = CellRichText('Add:', TextBlock(InlineFont(rFont='Times New Roman', u='single'), ' 7800 Northwestern Avenue, Racine, WI 53406'))
        rich_text_a9.font = Font(size=12, bold=True, name="Times New Roman")
        sheet["A9"] = rich_text_a9

        rich_text_a10 = CellRichText('Tel/Fax:', TextBlock(InlineFont(rFont='Times New Roman', u='single'), '800 6541704'))
        rich_text_a10.font = Font(size=12, bold=True, name="Times New Roman")
        sheet["A10"] = rich_text_a10   
        
        img = Image('logo.png')  # 替换为图片文件的路径

        # 设置图片的宽度和高度
        img.width = 200  # 设置宽度为100像素
        img.height = 100  # 设置高度为100像素
        
        sheet.add_image(img,f'B{picture_row + len(new_data) + 1}')
        
            
# 保存Excel文件
workbook.save('output_with_description.xlsx')